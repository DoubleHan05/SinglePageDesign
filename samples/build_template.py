"""生成给用户下载的空白 Excel 模板"""
from pathlib import Path
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.comments import Comment

HERE = Path(__file__).parent
out = HERE / "产品单页模板.xlsx"

wb = openpyxl.Workbook()

# ───────── Sheet1：产品数据 ─────────
ws1 = wb.active
ws1.title = "产品数据"

group_fill = PatternFill("solid", fgColor="EEF4FF")
group_font = Font(bold=True, color="2B6CFF")

def add_group(ws, title):
    ws.append([title, ""])
    row = ws.max_row
    ws.cell(row=row, column=1).font = group_font
    ws.cell(row=row, column=1).fill = group_fill
    ws.cell(row=row, column=2).fill = group_fill

def add(ws, k, v="", note=None):
    ws.append([k, v])
    if note:
        ws.cell(row=ws.max_row, column=1).comment = Comment(note, "系统")

add_group(ws1, "== 基础信息 ==")
add(ws1, "型号", "请输入产品型号", "产品型号，会显示在左上区域【产品型号：xxx】")

add_group(ws1, "== 图片（可直接在 B 列单元格插入图片） ==")
add(ws1, "产品图", "", "主产品图，可在此行的 B 列插入图片；也可填写图片文件路径")
add(ws1, "细节图1", "", "细节图第 1 张；可在 B 列插入图片")
add(ws1, "细节图2", "", "细节图第 2 张（可选）")

add_group(ws1, "== 产品特点（想加几条就加几行，编号自增） ==")
add(ws1, "特点1", "免安装驱动、异地远程打印、多人跨端打印、支持多个系统")
add(ws1, "特点2", "普通用户：番茄标签 + 快麦云打印机")
add(ws1, "特点3", "第三方软件用户：快麦云打印机无需开发支持所有软件")
add(ws1, "特点4", "企业开发者：支持 SaaS 对接、免费 API 接口")

add_group(ws1, "== 应用场景（值格式：圆圈内文字|下方标签） ==")
add(ws1, "场景1", "仓|仓储")
add(ws1, "场景2", "物|物流")
add(ws1, "场景3", "零|零售")

ws1.column_dimensions["A"].width = 34
ws1.column_dimensions["B"].width = 62

# ───────── Sheet2：参数表 ─────────
ws2 = wb.create_sheet("参数表")

header_font = Font(bold=True, color="FFFFFF", size=11)
header_fill = PatternFill("solid", fgColor="2B6CFF")
ws2.append(["参数名称", "参数值"])
for c in ws2[1]:
    c.font = header_font
    c.fill = header_fill
    c.alignment = Alignment(horizontal="center", vertical="center")

params = [
    ("电池寿命", ""),
    ("充放电次数", ""),
    ("指令集", "ESC\\POS、TSPL"),
    ("打印内容", "英文、中文简体、中文繁体、图片、字符、线条、符号、一维条码、二维码"),
    ("编码方式", "GBK、BIG5、KSC5601、UTF-8"),
    ("字体类型", ""),
    ("字符大小", ""),
    ("一维码", ""),
    ("二维码", ""),
    ("图形", ""),
    ("系统支持操作系统", "Windows、iOS、Android、Linux、MacOS"),
    ("对接平台/软件", "番茄标签、快递助手、风火递等"),
    ("固件升级方式", "支持 USB 本地升级、OTA 升级"),
    ("设备状态上报", "支持主动上报打印机状态（在线、离线）"),
    ("作业状态上报", "支持"),
    ("设备日志获取", "不支持"),
    ("切纸方式", "手动撕纸"),
    ("开盖方式", "上翻盖"),
    ("纸仓形式", "有纸仓"),
    ("切刀寿命", ""),
    ("出纸方式", "前出纸"),
    ("外观尺寸（长*宽*高）", "请填写"),
    ("包装尺寸（长*宽*高）", "请填写"),
    ("裸机重量（不包含电池/适配器）", ""),
    ("单台整机重量", ""),
    ("配件", "电源适配器、数据线、说明书（含保修卡）"),
    ("台数", ""),
    ("整箱重量", ""),
    ("整箱尺寸（长*宽*高）", ""),
    ("按键", ""),
    ("语音提示", ""),
    ("其他", ""),
    ("工作温度/湿度", "温度：0-45℃；湿度：20-90%"),
    ("存储温度/湿度", "温度：-10-60℃；湿度：10-90%"),
    ("纸张回退", "支持"),
    ("自动进纸", "支持"),
    ("压缩方式", ""),
    ("认证标准", "3C认证"),
    ("ESD要求", ""),
    ("防护等级", ""),
    ("目标市场/场景", "连锁餐饮、连锁商超、连锁门店"),
]
for k, v in params:
    ws2.append([k, v])

ws2.column_dimensions["A"].width = 34
ws2.column_dimensions["B"].width = 62

# ───────── Sheet3：使用说明 ─────────
ws3 = wb.create_sheet("使用说明")
lines = [
    "使用说明",
    "",
    "【Sheet1 - 产品数据】",
    "填写型号、图片路径（或直接在 B 列插入图片）、产品特点、应用场景。",
    "",
    "【Sheet2 - 参数表】",
    "直接填写参数名称和参数值，无需加任何前缀，方便从其他地方粘贴。",
    "想增减参数行，直接在表格末尾添加或删除行即可。",
    "",
    "【其他说明】",
    "· 产品特点：特点1、特点2... 想写几条加几行",
    "· 细节图：细节图1、细节图2... 支持多张，可直接在 B 列插入图片",
    "· 应用场景格式：'仓|仓储'（竖线前=圆圈内文字，竖线后=下方标签）",
    "· 参数值留空时，页面上显示 '/'",
    "· 填好后在工具网页拖入此文件，预览确认后下载图片。",
]
for i, line in enumerate(lines, start=1):
    ws3.cell(row=i, column=1, value=line)
ws3.column_dimensions["A"].width = 80

wb.save(out)
print("wrote", out)
