#!/usr/bin/env python3
"""
产品单页生成工具

用法:
    python generate.py <excel_path> [--out-dir <dir>] [--no-image]

Excel 结构：
    Sheet1「产品数据」（两列：参数名 / 参数值）
        型号、产品图、细节图1/2/...、特点1/2/...、场景1/2/...
    Sheet2「参数表」（两列：参数名称 / 参数值，第一行为表头）
        任意条目，直接填写即可
"""
import argparse
import os
import re
import sys
import zipfile
from html import escape
from pathlib import Path
from typing import Dict, List, Optional, Tuple
from xml.etree import ElementTree as ET

import openpyxl

HERE = Path(__file__).parent
TEMPLATE_PATH = HERE / "templates" / "template.html"

# 支持的图片扩展名
_IMG_EXTS = (".png", ".jpg", ".jpeg", ".gif", ".bmp", ".webp", ".svg")

# openpyxl / xlsx 命名空间
_NS = {
    "main": "http://schemas.openxmlformats.org/spreadsheetml/2006/main",
    "xdr":  "http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing",
    "a":    "http://schemas.openxmlformats.org/drawingml/2006/main",
    "r":    "http://schemas.openxmlformats.org/officeDocument/2006/relationships",
    "rel":  "http://schemas.openxmlformats.org/package/2006/relationships",
}


# ─────────────────────── 图片提取：浮动图片（openpyxl 可读） ───────────────────────

def _extract_floating_images(ws, out_dir: Path) -> Dict[int, List[Path]]:
    """把 sheet 中的浮动嵌入图片按行号提取出来。

    Returns:
        {row_1based: [saved_path, ...]}
    """
    out_dir.mkdir(parents=True, exist_ok=True)
    row_map: Dict[int, List[Path]] = {}
    images = getattr(ws, "_images", []) or []
    for idx, img in enumerate(images):
        row_1based = _anchor_row(img)
        if row_1based is None:
            continue
        blob = _read_image_blob(img)
        if blob is None:
            continue
        ext = _guess_image_ext(blob, getattr(img, "format", ""))
        fpath = out_dir / f"embed_{idx}_r{row_1based}{ext}"
        with open(fpath, "wb") as f:
            f.write(blob)
        row_map.setdefault(row_1based, []).append(fpath)
    return row_map


def _anchor_row(img) -> Optional[int]:
    """从 openpyxl 图片对象上读出锚点行号（1-based）。失败返回 None。"""
    try:
        anchor = img.anchor
        frm = getattr(anchor, "_from", None) or getattr(anchor, "from_", None)
        if frm is not None:
            return int(frm.row) + 1
    except AttributeError:
        return None
    return None


def _read_image_blob(img) -> Optional[bytes]:
    """兼容不同 openpyxl 版本读出图片二进制。"""
    try:
        reader = getattr(img, "_data", None)
        if callable(reader):
            return reader()
    except Exception:
        pass
    try:
        return img.ref.read()
    except Exception:
        return None


def _guess_image_ext(blob: bytes, fmt_hint: str = "") -> str:
    fmt = (fmt_hint or "").lower()
    if fmt in ("jpg", "jpeg"): return ".jpg"
    if fmt == "gif":  return ".gif"
    if fmt == "bmp":  return ".bmp"
    if fmt == "webp": return ".webp"
    if fmt == "png":  return ".png"
    # 从 magic header 猜
    if blob[:3] == b"\xff\xd8\xff":       return ".jpg"
    if blob[:4] == b"GIF8":               return ".gif"
    if blob[:8].startswith(b"\x89PNG"):   return ".png"
    return ".png"


# ─────────────────────── 图片提取：单元格内嵌图片（WPS/新版 Excel） ───────────────────────
# WPS / 新版 Excel 的"在单元格中放置图片"实际是 =DISPIMG("ID_xxx",1) 公式，
# 图片本体存在 xl/cellimages.xml + xl/media/。需要手动解 zip。

def _load_cellimages_rels(zf: zipfile.ZipFile) -> Dict[str, str]:
    """cellimages.xml.rels 中的 rId → 媒体文件路径（xl/media/xxx）"""
    if "xl/_rels/cellimages.xml.rels" not in zf.namelist():
        return {}
    root = ET.fromstring(zf.read("xl/_rels/cellimages.xml.rels"))
    result: Dict[str, str] = {}
    for rel in root.findall("rel:Relationship", _NS):
        rid = rel.get("Id"); target = rel.get("Target", "")
        if target.startswith("../"):
            target = "xl/" + target[3:]
        elif not target.startswith("xl/"):
            target = "xl/" + target
        if rid:
            result[rid] = target
    return result


def _load_cellimage_names(zf: zipfile.ZipFile, rid_to_media: Dict[str, str]) -> Dict[str, str]:
    """cellimages.xml 中的图片名（如 "ID_xxx"）→ 媒体文件路径"""
    if "xl/cellimages.xml" not in zf.namelist():
        return {}
    root = ET.fromstring(zf.read("xl/cellimages.xml"))
    result: Dict[str, str] = {}
    for pic in root.iter(f"{{{_NS['xdr']}}}pic"):
        nv = pic.find("xdr:nvPicPr/xdr:cNvPr", _NS)
        blip = pic.find("xdr:blipFill/a:blip", _NS)
        if nv is None or blip is None:
            continue
        pic_name = nv.get("name") or nv.get("descr")
        embed = blip.get(f"{{{_NS['r']}}}embed")
        if pic_name and embed and embed in rid_to_media:
            result[pic_name] = rid_to_media[embed]
    return result


def _find_dispimg_cells(zf: zipfile.ZipFile, sheet_index: int = 0) -> List[Tuple[int, str]]:
    """扫描指定 sheet.xml，返回 [(row_1based, picture_name), ...]"""
    sheet_paths = sorted(n for n in zf.namelist()
                         if n.startswith("xl/worksheets/sheet") and n.endswith(".xml"))
    if sheet_index >= len(sheet_paths):
        return []
    root = ET.fromstring(zf.read(sheet_paths[sheet_index]))
    out: List[Tuple[int, str]] = []
    for cell in root.iter(f"{{{_NS['main']}}}c"):
        f_el = cell.find("main:f", _NS)
        if f_el is None or not f_el.text:
            continue
        m = re.search(r'DISPIMG\(\s*"([^"]+)"', f_el.text)
        if not m:
            continue
        row_m = re.search(r"(\d+)$", cell.get("r", ""))
        if not row_m:
            continue
        out.append((int(row_m.group(1)), m.group(1)))
    return out


def _extract_cell_embedded_images(
    xlsx_path: Path,
    out_dir: Path,
    sheet_index: int = 0,
) -> Dict[int, List[Path]]:
    """
    解析 WPS / 新版 Excel 的"在单元格中放置图片"。

    Returns:
        {row_1based: [saved_path, ...]}
    """
    row_map: Dict[int, List[Path]] = {}
    if not xlsx_path.exists():
        return row_map
    try:
        with zipfile.ZipFile(xlsx_path) as zf:
            rid_to_media = _load_cellimages_rels(zf)
            if not rid_to_media:
                return row_map
            name_to_media = _load_cellimage_names(zf, rid_to_media)
            if not name_to_media:
                return row_map

            zf_names = set(zf.namelist())
            dispimg_cells = _find_dispimg_cells(zf, sheet_index)
            if not dispimg_cells:
                return row_map

            out_dir.mkdir(parents=True, exist_ok=True)
            for row_1, pic_name in dispimg_cells:
                media_rel = name_to_media.get(pic_name)
                if not media_rel or media_rel not in zf_names:
                    continue
                ext = os.path.splitext(media_rel)[1] or ".png"
                fpath = out_dir / f"cell_{pic_name}{ext}"
                with zf.open(media_rel) as src, open(fpath, "wb") as dst:
                    dst.write(src.read())
                row_map.setdefault(row_1, []).append(fpath)
    except (zipfile.BadZipFile, ET.ParseError, OSError) as e:
        print(f"⚠️  解析单元格嵌入图片失败：{e.__class__.__name__}: {e}")
    return row_map


# ─────────────────────── Excel 读取 ───────────────────────

def read_excel(path: Path, image_out_dir: Optional[Path] = None) -> List[Tuple[str, str]]:
    """
    读取 Excel。Sheet1 是产品数据，Sheet2 是参数表（无需加"参数."前缀）。
    若某一行第二列为空且该行嵌入了图片，则用图片路径作为值。

    Args:
        path: xlsx 文件路径
        image_out_dir: 把嵌入图片解出到该目录（默认 excel 所在目录/_embed）

    Returns:
        [(key, value), ...] 顺序为 sheet1 后接 sheet2
    """
    wb = openpyxl.load_workbook(path, data_only=True)

    if image_out_dir is None:
        image_out_dir = path.parent / "_embed"

    data: List[Tuple[str, str]] = []

    # ─── Sheet1：产品数据 ───
    ws1 = wb.worksheets[0]
    row_images = _extract_floating_images(ws1, image_out_dir)
    # 合并 WPS / 新版 Excel 的"单元格内嵌图片"
    for r, lst in _extract_cell_embedded_images(path, image_out_dir, 0).items():
        row_images.setdefault(r, []).extend(lst)

    img_key_re = re.compile(r"^(产品图|细节图\s*\d*)$")
    for row_idx, row in enumerate(ws1.iter_rows(values_only=True), start=1):
        if not row:
            continue
        key = row[0]; val = row[1] if len(row) > 1 else None
        if key is None:
            continue
        key = str(key).strip()
        if not key:
            continue
        val = "" if val is None else str(val).strip()

        # 图片列（产品图 / 细节图N）优先使用嵌入图片
        if img_key_re.match(key):
            imgs = row_images.get(row_idx, [])
            if imgs and not val:
                val = str(imgs[0])

        data.append((key, val))

    # ─── Sheet2：参数表（如果存在） ───
    if len(wb.worksheets) >= 2:
        ws2 = wb.worksheets[1]
        for row in ws2.iter_rows(values_only=True, min_row=2):
            if not row:
                continue
            key = row[0]; val = row[1] if len(row) > 1 else None
            if key is None:
                continue
            key = str(key).strip()
            if not key:
                continue
            val = "" if val is None else str(val).strip()
            data.append((f"参数.{key}", val))

    return data


# ─────────────────────── HTML 片段构造 ───────────────────────

def build_features(items: List[str]) -> str:
    if not items:
        return '<li style="list-style:none;color:#999">（未提供产品特点）</li>'
    return "\n".join(f"<li>{escape(x)}</li>" for x in items)


def build_scenes(items: List[str]) -> str:
    if not items:
        return ""
    parts = []
    for raw in items:
        ch, label = (raw.split("|", 1) if "|" in raw else (raw, raw))
        parts.append(
            f'<div class="scene"><div class="circle">{escape(ch.strip())}</div>'
            f'<div class="label">{escape(label.strip())}</div></div>'
        )
    return "\n".join(parts)


def build_param_rows(rows: List[Tuple[str, str]]) -> str:
    if not rows:
        return '<tr><td class="k">（无参数）</td><td class="v">/</td></tr>'
    out = []
    for k, v in rows:
        v = v if v else "/"
        out.append(f'<tr><td class="k">{escape(k)}</td><td class="v">{escape(v)}</td></tr>')
    return "\n".join(out)


def build_detail_imgs(srcs: List[str]) -> str:
    if not srcs:
        return '<div class="product-img" style="color:#bbb;font-size:12px">（未提供细节图）</div>'
    return "\n".join(
        f'<div class="product-img"><img src="{escape(src, quote=True)}" alt="detail"></div>'
        for src in srcs
    )


# ─────────────────────── 字段派发表 ───────────────────────

def _classify(key: str) -> Tuple[str, Optional[str]]:
    """
    把 Excel 里的 key 归类到内部字段。
    Returns: (category, slot_name)
        category: model | product_img | detail_img | feature | scene | param
        slot_name: 派发到 category 下的具体 slot（当前仅 model / product_img 有效）
    """
    if key == "型号":
        return ("model", "MODEL")
    if key == "产品图":
        return ("product_img", "PRODUCT_IMG")
    if re.match(r"^细节图\s*\d*$", key):
        return ("detail_img", None)
    if re.match(r"^特点\s*\d+$", key):
        return ("feature", None)
    if re.match(r"^场景\s*\d+$", key):
        return ("scene", None)
    # 参数.xxx 或未识别的 key 都当参数处理
    return ("param", None)


def render(data: List[Tuple[str, str]], template: str, theme: str = "blue") -> str:
    if theme not in ("blue", "light"):
        theme = "blue"
    scalars: Dict[str, str] = {"MODEL": "", "PRODUCT_IMG": "", "THEME": theme}
    features: List[str] = []
    scenes: List[str] = []
    detail_imgs: List[str] = []
    params: List[Tuple[str, str]] = []

    for key, val in data:
        cat, slot = _classify(key)
        if cat in ("model", "product_img") and slot:
            scalars[slot] = val
        elif cat == "feature" and val:
            features.append(val)
        elif cat == "scene" and val:
            scenes.append(val)
        elif cat == "detail_img" and val:
            detail_imgs.append(val)
        elif cat == "param":
            pname = key[len("参数."):].strip() if key.startswith("参数.") else key
            if pname:
                params.append((pname, val))

    # 所有插槽都做 HTML 转义（属性值也安全，斜杠不受影响）
    html = template
    for k, v in scalars.items():
        html = html.replace("{{" + k + "}}", escape(v, quote=True))
    html = html.replace("{{FEATURES}}", build_features(features))
    html = html.replace("{{SCENES}}", build_scenes(scenes))
    html = html.replace("{{PARAM_ROWS}}", build_param_rows(params))
    html = html.replace("{{DETAIL_IMGS}}", build_detail_imgs(detail_imgs))
    return html


# ─────────────────────── HTML → PNG ───────────────────────

def html_to_png(html_path: Path, png_path: Path) -> bool:
    """优先使用 Playwright（渲染最接近浏览器），退回到 WeasyPrint + PyMuPDF。"""
    if _render_with_playwright(html_path, png_path):
        return True
    return _render_with_weasyprint(html_path, png_path)


def _render_with_playwright(html_path: Path, png_path: Path) -> bool:
    try:
        from playwright.sync_api import sync_playwright  # type: ignore
    except ImportError:
        return False
    last_err: Optional[Exception] = None
    # 依次尝试：系统 Chrome → 系统 Edge → 自带 chromium
    for launcher in (
        lambda p: p.chromium.launch(channel="chrome"),
        lambda p: p.chromium.launch(channel="msedge"),
        lambda p: p.chromium.launch(),
    ):
        try:
            with sync_playwright() as p:
                browser = launcher(p)
                page = browser.new_page(viewport={"width": 1062, "height": 800},
                                        device_scale_factor=2)
                page.goto(html_path.absolute().as_uri())
                page.wait_for_load_state("networkidle")
                page.screenshot(path=str(png_path), full_page=True)
                browser.close()
            return True
        except Exception as e:  # noqa: BLE001 — 记录以便切换到下一个 channel
            last_err = e
            continue
    if last_err:
        print(f"ℹ️  Playwright 不可用（{last_err.__class__.__name__}），改用 WeasyPrint。")
    return False


def _render_with_weasyprint(html_path: Path, png_path: Path) -> bool:
    try:
        from weasyprint import HTML  # type: ignore
        import fitz  # type: ignore
    except ImportError:
        print("⚠️  未安装 weasyprint / pymupdf，无法生成图片。")
        return False
    try:
        pdf_bytes = HTML(filename=str(html_path)).write_pdf()
        doc = fitz.open(stream=pdf_bytes, filetype="pdf")
        import tempfile as _tf
        tmp_png = Path(_tf.mkstemp(suffix=".png")[1])

        zoom = 2.0
        mat = fitz.Matrix(zoom, zoom)
        pixmaps = [page.get_pixmap(matrix=mat, alpha=False) for page in doc]

        if len(pixmaps) == 1:
            pixmaps[0].save(str(tmp_png))
            _trim_bottom_whitespace(tmp_png, png_path)
        else:
            _stack_pixmaps_vertically(pixmaps, tmp_png)
            _copy_file(tmp_png, png_path)
        doc.close()
        return True
    except Exception as e:
        print(f"⚠️  图片生成失败：{e.__class__.__name__}: {e}")
        return False


def _trim_bottom_whitespace(src_png: Path, dst_png: Path) -> None:
    """把底部大片浅色背景裁掉。失败时直接复制。"""
    try:
        from PIL import Image, ImageChops
        im = Image.open(str(src_png)).convert("RGB")
        bg = Image.new("RGB", im.size, (244, 246, 248))
        diff = ImageChops.difference(im, bg)
        bbox = diff.getbbox()
        if bbox:
            pad = 20
            im = im.crop((0, 0, im.width, min(im.height, bbox[3] + pad)))
        im.save(str(dst_png))
    except Exception:
        _copy_file(src_png, dst_png)


def _stack_pixmaps_vertically(pixmaps, dst_png: Path) -> None:
    from PIL import Image
    import io
    imgs = [Image.open(io.BytesIO(pm.tobytes("png"))) for pm in pixmaps]
    w = max(i.width for i in imgs)
    h = sum(i.height for i in imgs)
    combo = Image.new("RGB", (w, h), "white")
    y = 0
    for im in imgs:
        combo.paste(im, (0, y)); y += im.height
    combo.save(str(dst_png))


def _copy_file(src: Path, dst: Path) -> None:
    with open(src, "rb") as f, open(dst, "wb") as g:
        g.write(f.read())


# ─────────────────────── CLI 入口 ───────────────────────

def main() -> None:
    ap = argparse.ArgumentParser()
    ap.add_argument("excel", help="输入 Excel 文件路径")
    ap.add_argument("--out-dir", default=str(HERE / "output"),
                    help="输出目录（默认 ./output）")
    ap.add_argument("--no-image", action="store_true", help="仅生成 HTML，不出图")
    args = ap.parse_args()

    excel_path = Path(args.excel).expanduser().resolve()
    if not excel_path.exists():
        print(f"❌ Excel 不存在：{excel_path}"); sys.exit(1)

    out_dir = Path(args.out_dir); out_dir.mkdir(parents=True, exist_ok=True)
    template = TEMPLATE_PATH.read_text(encoding="utf-8")
    data = read_excel(excel_path)
    html = render(data, template)

    stem = excel_path.stem
    html_path = out_dir / f"{stem}.html"
    html_path.write_text(html, encoding="utf-8")
    print(f"✅ HTML: {html_path}")

    if not args.no_image:
        png_path = out_dir / f"{stem}.png"
        if html_to_png(html_path, png_path):
            print(f"✅ PNG:  {png_path}")


if __name__ == "__main__":
    main()
